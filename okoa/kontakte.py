"""Externe Kontakte sammeln und als Excel ausgeben.

Anders als der Rest des Projekts ist das Ergebnis hier eine **personenbezogene
Liste**, keine Kennzahl.  Sie hat einen eigenen Zweck (Adressbestand,
Lieferantenuebersicht) und faellt damit unter eine andere datenschutzrechtliche
Bewertung als die aggregierte Auswertung -- siehe docs/10-kontaktliste.md.

Die Firmenzuordnung kennt drei Herkuenfte, die in der Ausgabe immer sichtbar
sind.  Wer eine Zahl weiterverwendet, soll wissen, wie sicher sie ist:

    signatur   mehrfach uebereinstimmend aus Signaturen gelesen  -- belastbar
    domain     aus dem Domainnamen abgeleitet                    -- Lesehilfe
    (leer)     nichts Eindeutiges gefunden                       -- ehrlich
"""

from __future__ import annotations

from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from .model import EXTERN, RICHTUNG_GESENDET, Nachricht, Vorgang
from .normalize import adresse_normalisieren, domain_von, ist_automat
from .signaturen import aus_domain, konsens


HERKUNFT_SIGNATUR = "Signatur"
HERKUNFT_DOMAIN = "Domainname"
HERKUNFT_KEINE = ""

SPALTEN = [
    "E-Mail", "Anzeigename", "Funktion", "Telefon", "Mobil",
    "Domain", "Unternehmen", "Herkunft Unternehmen", "Belege Unternehmen",
    "Signaturbelege", "Kategorie", "Nachrichten", "gesendet", "empfangen", "Vorgänge",
    "Erstkontakt", "Letzter Kontakt", "Letzte eigene Nachricht",
    "Letzte Nachricht von dort", "Tage seit letztem Kontakt", "Status",
]

# Spalten, die als echter Zeitstempel geschrieben werden -- nicht als Text.
# Sonst sortiert Excel nach Zeichenkette und stellt den 01.02. vor den 30.01.
ZEITSPALTEN = ["Erstkontakt", "Letzter Kontakt", "Letzte eigene Nachricht",
               "Letzte Nachricht von dort"]
ZEITFORMAT = "DD.MM.YYYY HH:MM"

# Ab wann ein Kontakt als eingeschlafen gilt.  Reine Lesehilfe, keine Bewertung.
INAKTIV_AB_TAGEN = 180


# Spalten, die als Text erzwungen werden -- sonst frisst Excel die fuehrende
# Null einer Vorwahl und macht aus '0170 1234567' eine Zahl.
TEXTSPALTEN = ["Telefon", "Mobil"]


@dataclass
class Beleg:
    """Was die Extraktion je Nachricht ueber einen Kontakt gesehen hat.

    Firma gilt fuer das ganze Haus, Funktion und Rufnummern gelten nur fuer
    diese eine Person -- deshalb werden sie getrennt zusammengefasst.
    """

    adresse: str
    anzeigename: str = ""
    firma_kandidat: str | None = None
    funktion_kandidat: str | None = None
    telefon_kandidat: str | None = None
    mobil_kandidat: str | None = None


@dataclass
class Kontakt:
    adresse: str
    domain: str
    anzeigenamen: Counter = field(default_factory=Counter)
    firma_kandidaten: list[str] = field(default_factory=list)
    funktion_kandidaten: list[str] = field(default_factory=list)
    telefon_kandidaten: list[str] = field(default_factory=list)
    mobil_kandidaten: list[str] = field(default_factory=list)
    signaturen_gesehen: int = 0
    nachrichten: int = 0
    gesendet: int = 0
    empfangen: int = 0
    vorgaenge: set[str] = field(default_factory=set)
    erstkontakt: datetime | None = None
    letzter_kontakt: datetime | None = None
    letzte_eigene: datetime | None = None
    letzte_fremde: datetime | None = None

    def anzeigename(self) -> str:
        """Der haeufigste Anzeigename -- Schreibweisen schwanken."""
        return self.anzeigenamen.most_common(1)[0][0] if self.anzeigenamen else ""


def sammeln(vorgaenge: list[Vorgang], belege: dict[str, list[Beleg]] | None = None,
            automaten_ausschliessen: bool = True) -> list[Kontakt]:
    """Baut die Kontaktliste aus den bereits klassifizierten Nachrichten.

    Ohne Belege (also aus der Zwischendatei heraus) entsteht die Liste trotzdem
    -- dann eben ohne Anzeigenamen und ohne Signaturauswertung.
    """
    belege = belege or {}
    kontakte: dict[str, Kontakt] = {}

    for vorgang in vorgaenge:
        for nachricht in vorgang.nachrichten:
            if not nachricht.ist_auswertbar:
                continue
            paare = zip([nachricht.absender_id, *nachricht.empfaenger_ids],
                        [nachricht.absender_klasse, *nachricht.empfaenger_klassen])
            for adresse, klasse in paare:
                if klasse != EXTERN:
                    continue
                adresse = adresse_normalisieren(adresse)
                if not adresse or (automaten_ausschliessen and ist_automat(adresse)):
                    continue

                kontakt = kontakte.get(adresse)
                if kontakt is None:
                    kontakt = Kontakt(adresse=adresse, domain=domain_von(adresse))
                    kontakte[adresse] = kontakt
                kontakt.nachrichten += 1
                if nachricht.richtung == RICHTUNG_GESENDET:
                    kontakt.gesendet += 1
                    if (kontakt.letzte_eigene is None
                            or nachricht.zeitstempel > kontakt.letzte_eigene):
                        kontakt.letzte_eigene = nachricht.zeitstempel
                else:
                    kontakt.empfangen += 1
                    if (kontakt.letzte_fremde is None
                            or nachricht.zeitstempel > kontakt.letzte_fremde):
                        kontakt.letzte_fremde = nachricht.zeitstempel
                kontakt.vorgaenge.add(vorgang.thread_id)
                if kontakt.erstkontakt is None or nachricht.zeitstempel < kontakt.erstkontakt:
                    kontakt.erstkontakt = nachricht.zeitstempel
                if kontakt.letzter_kontakt is None or nachricht.zeitstempel > kontakt.letzter_kontakt:
                    kontakt.letzter_kontakt = nachricht.zeitstempel

    for adresse, eintraege in belege.items():
        kontakt = kontakte.get(adresse_normalisieren(adresse))
        if kontakt is None:
            continue
        for beleg in eintraege:
            if beleg.anzeigename:
                kontakt.anzeigenamen[beleg.anzeigename.strip()] += 1
            if beleg.firma_kandidat:
                kontakt.firma_kandidaten.append(beleg.firma_kandidat)
            if beleg.funktion_kandidat:
                kontakt.funktion_kandidaten.append(beleg.funktion_kandidat)
            if beleg.telefon_kandidat:
                kontakt.telefon_kandidaten.append(beleg.telefon_kandidat)
            if beleg.mobil_kandidat:
                kontakt.mobil_kandidaten.append(beleg.mobil_kandidat)
            if any((beleg.firma_kandidat, beleg.funktion_kandidat,
                    beleg.telefon_kandidat, beleg.mobil_kandidat)):
                kontakt.signaturen_gesehen += 1

    return sorted(kontakte.values(), key=lambda k: (-k.nachrichten, k.adresse))


def firmen_je_domain(kontakte: list[Kontakt]) -> dict[str, tuple[str, int]]:
    """Konsens ueber alle Kontakte einer Domain.

    Bewusst auf Domainebene: Signaturen einzelner Personen sind lueckenhaft,
    aber eine Firmierung gilt fuer das ganze Haus.  So profitiert auch der
    Kollege, der nie eine Signatur mitgeschickt hat.
    """
    je_domain: dict[str, list[str]] = {}
    for kontakt in kontakte:
        je_domain.setdefault(kontakt.domain, []).extend(kontakt.firma_kandidaten)
    return {domain: konsens(kandidaten) for domain, kandidaten in je_domain.items()}


def als_zeilen(kontakte: list[Kontakt], kategorien: dict[str, str] | None = None,
               stichtag: datetime | None = None) -> list[dict]:
    kategorien = kategorien or {}
    stichtag = stichtag or datetime.now()
    firmen = firmen_je_domain(kontakte)

    zeilen = []
    for kontakt in kontakte:
        name, belege = firmen.get(kontakt.domain, (None, 0))
        if name:
            unternehmen, herkunft = name, HERKUNFT_SIGNATUR
        else:
            unternehmen, herkunft = aus_domain(kontakt.domain), HERKUNFT_DOMAIN
            belege = 0
        # Nie negativ: ein Kontakt von heute Nachmittag liegt sonst "-1 Tage"
        # zurueck, was in der Tabelle wie ein Fehler aussieht.
        # Funktion und Rufnummern gelten nur fuer diese eine Person -- der
        # Konsens laeuft deshalb je Adresse, nicht je Domain.
        funktion, _ = konsens(kontakt.funktion_kandidaten)
        telefon, _ = konsens(kontakt.telefon_kandidaten)
        mobil, _ = konsens(kontakt.mobil_kandidaten)

        tage = (max(0, (stichtag - kontakt.letzter_kontakt).days)
                if kontakt.letzter_kontakt else None)
        zeilen.append({
            "E-Mail": kontakt.adresse,
            "Anzeigename": kontakt.anzeigename(),
            "Funktion": funktion or "",
            "Telefon": telefon or "",
            "Mobil": mobil or "",
            "Domain": kontakt.domain,
            "Unternehmen": unternehmen,
            "Herkunft Unternehmen": herkunft,
            "Belege Unternehmen": belege,
            "Signaturbelege": kontakt.signaturen_gesehen,
            "Kategorie": kategorien.get(kontakt.domain, ""),
            "Nachrichten": kontakt.nachrichten,
            "gesendet": kontakt.gesendet,
            "empfangen": kontakt.empfangen,
            "Vorgänge": len(kontakt.vorgaenge),
            # Echte Zeitstempel, damit sich die Liste in Excel sinnvoll
            # sortieren und filtern laesst.
            "Erstkontakt": kontakt.erstkontakt or "",
            "Letzter Kontakt": kontakt.letzter_kontakt or "",
            # Getrennt, weil die Frage "ist der Kontakt noch aktuell" anders
            # ausfaellt, je nachdem wer zuletzt geschrieben hat: eine eigene
            # Nachricht ohne Antwort ist etwas anderes als ein laufender Dialog.
            "Letzte eigene Nachricht": kontakt.letzte_eigene or "",
            "Letzte Nachricht von dort": kontakt.letzte_fremde or "",
            "Tage seit letztem Kontakt": tage if tage is not None else "",
            "Status": ("aktiv" if tage is not None and tage <= INAKTIV_AB_TAGEN
                       else "eingeschlafen"),
        })
    return zeilen


def schreiben(zeilen: list[dict], pfad: Path | str) -> Path:
    """Schreibt als .xlsx mit Autofilter; ohne openpyxl als .csv."""
    from . import mapping

    pfad = Path(pfad)
    if pfad.suffix.lower() != ".xlsx":
        # In der CSV gibt es keine Zellformate -- dort wird ausgeschrieben.
        zeilen = [{k: (v.strftime("%d.%m.%Y %H:%M") if isinstance(v, datetime) else v)
                   for k, v in zeile.items()} for zeile in zeilen]
    ziel = mapping.schreiben(zeilen, SPALTEN, pfad)
    if ziel.suffix.lower() == ".csv":
        zeilen = [{k: (v.strftime("%d.%m.%Y %H:%M") if isinstance(v, datetime) else v)
                   for k, v in zeile.items()} for zeile in zeilen]
        mapping.schreiben(zeilen, SPALTEN, ziel)
    if ziel.suffix.lower() == ".xlsx":
        _verschoenern(ziel, len(zeilen))
    return ziel


def _verschoenern(pfad: Path, anzahl: int) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return
    mappe = load_workbook(pfad)
    blatt = mappe.active
    blatt.auto_filter.ref = f"A1:{blatt.cell(row=1, column=len(SPALTEN)).column_letter}" \
                            f"{anzahl + 1}"
    for name, breite in (("E-Mail", 38), ("Anzeigename", 24), ("Funktion", 30),
                         ("Telefon", 20), ("Mobil", 20), ("Domain", 24),
                         ("Unternehmen", 34), ("Herkunft Unternehmen", 20)):
        index = SPALTEN.index(name) + 1
        blatt.column_dimensions[blatt.cell(row=1, column=index).column_letter].width = breite

    # Rufnummern als Text, sonst verschwindet die fuehrende Null der Vorwahl.
    for name in TEXTSPALTEN:
        index = SPALTEN.index(name) + 1
        for zeile in range(2, anzahl + 2):
            blatt.cell(row=zeile, column=index).number_format = "@"

    for name in ZEITSPALTEN:
        index = SPALTEN.index(name) + 1
        buchstabe = blatt.cell(row=1, column=index).column_letter
        blatt.column_dimensions[buchstabe].width = 19
        for zeile in range(2, anzahl + 2):
            blatt.cell(row=zeile, column=index).number_format = ZEITFORMAT
    blatt.freeze_panes = "A2"
    mappe.save(pfad)


def zusammenfassung(zeilen: list[dict]) -> dict:
    domains = {z["Domain"] for z in zeilen}
    mit_signatur = sum(1 for z in zeilen if z["Herkunft Unternehmen"] == HERKUNFT_SIGNATUR)
    return {
        "kontakte": len(zeilen),
        "domains": len(domains),
        "aus_signatur": mit_signatur,
        "anteil_aus_signatur": mit_signatur / len(zeilen) if zeilen else 0.0,
        "aktiv": sum(1 for z in zeilen if z["Status"] == "aktiv"),
        "mit_funktion": sum(1 for z in zeilen if z["Funktion"]),
        "mit_telefon": sum(1 for z in zeilen if z["Telefon"] or z["Mobil"]),
    }
