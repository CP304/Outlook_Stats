"""Mapping-Dateien fuer Fachbereiche und Domainkategorien.

Beide werden beim ersten Lauf automatisch erzeugt und nach Volumen absteigend
sortiert.  Das ist der eigentliche Setup-Trick: in der Praxis decken 15 bis 25
gepflegte Zeilen rund 80 % des Volumens ab.  Die Spalte 'Anteil kumuliert'
zeigt, wann man aufhoeren kann zu pflegen.

Vorhandene Dateien werden nie ueberschrieben -- neue Personen werden angehaengt.
"""

from __future__ import annotations

import csv
from collections import Counter
from pathlib import Path

from .model import EXTERN, INTERN, KONZERN, Nachricht, Vorgang
from .normalize import domain_von, ist_funktionspostfach


SPALTEN_PERSONEN = ["E-Mail", "Anzeigename", "Vorgaenge", "Nachrichten",
                    "Anteil kumuliert", "Fachbereich", "Rolle"]
SPALTEN_DOMAINS = ["Domain", "Vorgaenge", "Nachrichten", "Anteil kumuliert", "Kategorie"]

KATEGORIE_VORSCHLAEGE = ["Lieferant", "Dienstleister", "Kunde", "Behörde",
                         "Konzern/verbunden", "Sonstiges"]


def _kumuliert(zeilen: list[dict], schluessel: str) -> None:
    gesamt = sum(z[schluessel] for z in zeilen) or 1
    summe = 0
    for z in zeilen:
        summe += z[schluessel]
        z["Anteil kumuliert"] = round(summe / gesamt, 4)


def personen_sammeln(vorgaenge: list[Vorgang], nachrichten: list[Nachricht]) -> list[dict]:
    """Interne Personen nach Volumen, absteigend."""
    nachrichten_je: Counter = Counter()
    vorgaenge_je: dict[str, set[str]] = {}
    for v in vorgaenge:
        for n in v.nachrichten:
            if not n.ist_auswertbar:
                continue
            for adr, kl in zip([n.absender_id, *n.empfaenger_ids],
                               [n.absender_klasse, *n.empfaenger_klassen]):
                if kl not in (INTERN, KONZERN):
                    continue
                nachrichten_je[adr] += 1
                vorgaenge_je.setdefault(adr, set()).add(v.thread_id)

    zeilen = [
        {
            "E-Mail": adr,
            "Anzeigename": "",
            "Vorgaenge": len(vorgaenge_je.get(adr, ())),
            "Nachrichten": anzahl,
            "Anteil kumuliert": 0.0,
            # Funktionspostfaecher sind keine Personen und werden vorbelegt,
            # damit sie nicht als hyperaktive Kollegen erscheinen.
            "Fachbereich": "Funktionspostfach" if ist_funktionspostfach(adr) else "",
            "Rolle": "",
        }
        for adr, anzahl in nachrichten_je.most_common()
    ]
    _kumuliert(zeilen, "Nachrichten")
    return zeilen


def domains_sammeln(vorgaenge: list[Vorgang], nachrichten: list[Nachricht]) -> list[dict]:
    nachrichten_je: Counter = Counter()
    vorgaenge_je: dict[str, set[str]] = {}
    for v in vorgaenge:
        for n in v.nachrichten:
            if not n.ist_auswertbar:
                continue
            for adr, kl in zip([n.absender_id, *n.empfaenger_ids],
                               [n.absender_klasse, *n.empfaenger_klassen]):
                if kl != EXTERN:
                    continue
                d = domain_von(adr)
                nachrichten_je[d] += 1
                vorgaenge_je.setdefault(d, set()).add(v.thread_id)

    zeilen = [
        {
            "Domain": d,
            "Vorgaenge": len(vorgaenge_je.get(d, ())),
            "Nachrichten": anzahl,
            "Anteil kumuliert": 0.0,
            "Kategorie": "",
        }
        for d, anzahl in nachrichten_je.most_common()
    ]
    _kumuliert(zeilen, "Nachrichten")
    return zeilen


# ------------------------------------------------------------ schreiben

def _als_csv_schreiben(zeilen: list[dict], spalten: list[str], pfad: Path) -> None:
    with pfad.open("w", encoding="utf-8-sig", newline="") as f:
        schreiber = csv.DictWriter(f, fieldnames=spalten, delimiter=";")
        schreiber.writeheader()
        schreiber.writerows(zeilen)


def _als_excel_schreiben(zeilen: list[dict], spalten: list[str], pfad: Path,
                         auswahl: tuple[str, list[str]] | None = None) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        return False

    mappe = Workbook()
    blatt = mappe.active
    blatt.title = "Zuordnung"
    blatt.append(spalten)
    for zelle in blatt[1]:
        zelle.font = Font(bold=True)
        zelle.fill = PatternFill("solid", fgColor="DDE5F0")
    for z in zeilen:
        blatt.append([z.get(s, "") for s in spalten])

    if auswahl and zeilen:
        spaltenname, werte = auswahl
        index = spalten.index(spaltenname) + 1
        buchstabe = blatt.cell(row=1, column=index).column_letter
        pruefung = DataValidation(type="list", formula1='"' + ",".join(werte) + '"',
                                  allow_blank=True)
        blatt.add_data_validation(pruefung)
        pruefung.add(f"{buchstabe}2:{buchstabe}{len(zeilen) + 1}")

    for i, name in enumerate(spalten, start=1):
        breite = max(len(name) + 2, 14)
        if name in ("E-Mail", "Domain"):
            breite = 34
        blatt.column_dimensions[blatt.cell(row=1, column=i).column_letter].width = breite
    blatt.freeze_panes = "A2"
    mappe.save(pfad)
    return True


def schreiben(zeilen: list[dict], spalten: list[str], pfad: Path,
              auswahl: tuple[str, list[str]] | None = None) -> Path:
    """Schreibt als .xlsx, wenn openpyxl vorhanden ist, sonst als .csv."""
    from . import dateien

    pfad = Path(pfad)
    pfad.parent.mkdir(parents=True, exist_ok=True)

    def einmal(ziel: Path) -> None:
        if ziel.suffix.lower() == ".xlsx" and _als_excel_schreiben(
                zeilen, spalten, ziel, auswahl):
            return
        _als_csv_schreiben(zeilen, spalten, ziel.with_suffix(".csv"))

    ergebnis = dateien.mit_ausweichen(pfad, einmal)
    # Ohne openpyxl entsteht die .csv -- der zurueckgegebene Pfad muss das sagen.
    if ergebnis.suffix.lower() == ".xlsx" and not ergebnis.exists():
        return ergebnis.with_suffix(".csv")
    return ergebnis


def ergaenzen(neue_zeilen: list[dict], spalten: list[str], pfad: Path,
              schluessel: str) -> tuple[Path, int]:
    """Haengt unbekannte Eintraege an eine bestehende Datei an.

    Gepflegte Zuordnungen bleiben unangetastet -- niemand soll seine Arbeit
    beim zweiten Lauf verlieren.
    """
    pfad = Path(pfad)
    vorhanden = lesen(pfad)
    if not vorhanden:
        return schreiben(neue_zeilen, spalten, pfad), len(neue_zeilen)

    bekannt = {str(z.get(schluessel, "")).strip().lower() for z in vorhanden}
    ergaenzt = [z for z in neue_zeilen
                if str(z.get(schluessel, "")).strip().lower() not in bekannt]
    if not ergaenzt:
        return pfad, 0
    zusammen = vorhanden + ergaenzt
    ziel = pfad if pfad.exists() else pfad.with_suffix(".csv")
    return schreiben(zusammen, spalten, ziel), len(ergaenzt)


# ---------------------------------------------------------------- lesen

def lesen(pfad: Path) -> list[dict]:
    """Liest .xlsx oder .csv.  Fehlt die Datei, ist das Ergebnis leer."""
    pfad = Path(pfad)
    kandidaten = [pfad, pfad.with_suffix(".xlsx"), pfad.with_suffix(".csv")]
    for kandidat in kandidaten:
        if not kandidat.exists():
            continue
        if kandidat.suffix.lower() == ".xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError:
                continue
            mappe = load_workbook(kandidat, read_only=True, data_only=True)
            blatt = mappe.active
            reihen = blatt.iter_rows(values_only=True)
            kopf = [str(x) if x is not None else "" for x in next(reihen, [])]
            return [dict(zip(kopf, ["" if x is None else x for x in reihe]))
                    for reihe in reihen if any(x is not None for x in reihe)]
        with kandidat.open(encoding="utf-8-sig", newline="") as f:
            probe = f.read(2048)
            f.seek(0)
            trenner = ";" if probe.count(";") >= probe.count(",") else ","
            return list(csv.DictReader(f, delimiter=trenner))
    return []


def zuordnung_lesen(pfad: Path, schluessel: str, wert: str) -> dict[str, str]:
    """Ergibt {Adresse bzw. Domain: Label} -- leere Werte werden uebergangen."""
    ergebnis = {}
    for zeile in lesen(pfad):
        k = str(zeile.get(schluessel, "")).strip().lower()
        v = str(zeile.get(wert, "")).strip()
        if k and v:
            ergebnis[k] = v
    return ergebnis
